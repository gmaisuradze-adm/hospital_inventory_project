# Version: 1.3.4 - 2025-05-28 - Copilot Edit
# - Added PermissionRequiredMixin to EquipmentListView and EquipmentDetailView.
# - Ensured other views using PermissionRequiredMixin have appropriate permissions.

from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView, RedirectView
from django.views.generic.edit import FormView
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin
from django.contrib.messages.views import SuccessMessageMixin
from django.contrib import messages
from django.utils import timezone
from django.http import HttpResponse, JsonResponse
import csv
import json
from django.template.loader import render_to_string
from django.db.models import Q

# Import for Excel export
try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# Import for PDF export
try:
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

from .models import Equipment, Category, Status, Location, Supplier, EquipmentLog, DecommissionLog
from .forms import (
    EquipmentForm, CategoryForm, StatusForm, LocationForm, SupplierForm,
    EquipmentMarkForWriteOffForm, DecommissionConfirmForm, EquipmentRestoreForm
)
from django.utils.translation import gettext_lazy as _
from django.db.models import Q

# Helper function to get specific statuses
def get_status_instance(status_name_key, is_decommissioned_flag=None, is_marked_flag=None):
    query_params = {}
    if is_decommissioned_flag is not None:
        query_params['is_decommissioned'] = is_decommissioned_flag
    if is_marked_flag is not None:
        query_params['is_marked_for_write_off'] = is_marked_flag

    if not query_params:
        if status_name_key == 'Marked for Write-Off':
            query_params['is_marked_for_write_off'] = True
            query_params['is_decommissioned'] = False
        elif status_name_key == 'Decommissioned':
            query_params['is_decommissioned'] = True
        else:
            query_params['name__iexact'] = _(status_name_key)
            
    try:
        if not query_params:
            print(f"CRITICAL: Status lookup failed for key '{status_name_key}' and flags. No query parameters were set.")
            return None
        return Status.objects.get(**query_params)
    except Status.DoesNotExist:
        print(f"CRITICAL: Status with params {query_params} not found for key '{status_name_key}'!")
        return None
    except Status.MultipleObjectsReturned:
        print(f"CRITICAL: Multiple statuses found for params {query_params} for key '{status_name_key}'! Returning the first one.")
        return Status.objects.filter(**query_params).first()


# Redirect Views
class InventoryDashboardRedirectView(LoginRequiredMixin, RedirectView):
    """Redirect inventory dashboard requests to the main dashboard."""
    permanent = False
    pattern_name = 'core:dashboard'


# Equipment Views
class EquipmentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView): # Added PermissionRequiredMixin
    model = Equipment
    template_name = 'inventory/equipment_list.html'
    context_object_name = 'equipment_list'
    paginate_by = 15
    permission_required = 'inventory.view_equipment' # Added permission

    def get(self, request, *args, **kwargs):
        # Handle export requests
        export_format = request.GET.get('export')
        if export_format in ['csv', 'excel', 'pdf']:
            return self.export_data(export_format)
        return super().get(request, *args, **kwargs)

    def export_data(self, format_type):
        """Export equipment data in specified format with customization options"""
        queryset = self.get_queryset()
        
        # Handle selected IDs if provided
        selected_ids = self.request.GET.get('selected_ids')
        if selected_ids:
            try:
                ids = [int(id_str) for id_str in selected_ids.split(',')]
                queryset = queryset.filter(pk__in=ids)
            except (ValueError, TypeError):
                pass

        # Handle custom field selection
        fields = self.request.GET.get('fields', '').split(',')
        if fields and fields[0]:  # If fields parameter is provided and not empty
            valid_fields = {
                'asset_tag': 'asset_tag',
                'name': 'name',
                'category': 'category__name',
                'status': 'status__name',
                'location': 'current_location__name',
                'purchase_date': 'purchase_date',
                'purchase_cost': 'purchase_cost',
                'assigned_to': 'assigned_to__get_full_name',
            }
            selected_fields = [valid_fields[f] for f in fields if f in valid_fields]
            queryset = queryset.values(*selected_fields)

        # Handle sorting
        sort_by = self.request.GET.get('sort_by')
        if sort_by in valid_fields:
            queryset = queryset.order_by(valid_fields[sort_by])

        if format_type == 'csv':
            return self.export_csv(queryset, fields if fields and fields[0] else None)
        elif format_type == 'excel':
            return self.export_excel(queryset, fields if fields and fields[0] else None)
        elif format_type == 'pdf':
            return self.export_pdf(queryset, fields if fields and fields[0] else None)

    def export_csv(self, queryset, selected_fields=None):
        """Export to CSV format with field selection"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="equipment_list.csv"'
        
        writer = csv.writer(response)
        
        # Define field mappings for headers
        field_headers = {
            'asset_tag': _('ინვენტარის ნომერი'),
            'name': _('დასახელება'),
            'category': _('კატეგორია'),
            'status': _('სტატუსი'),
            'location': _('მდებარეობა'),
            'purchase_date': _('შეძენის თარიღი'),
            'purchase_cost': _('შეძენის ფასი'),
            'assigned_to': _('მინიჭებული'),
        }
        
        # Write headers
        if selected_fields:
            headers = [field_headers[f] for f in selected_fields]
        else:
            headers = list(field_headers.values())
        writer.writerow(headers)
        
        # Write data rows
        for item in queryset:
            if isinstance(item, dict):  # For custom field selection
                row = [item.get(f, '') for f in selected_fields] if selected_fields else list(item.values())
            else:  # For full equipment objects
                if selected_fields:
                    row = []
                    for field in selected_fields:
                        if field == 'category':
                            row.append(item.category.name if item.category else '')
                        elif field == 'status':
                            row.append(item.status.name if item.status else '')
                        elif field == 'location':
                            row.append(item.current_location.get_full_path() if item.current_location else '')
                        elif field == 'assigned_to':
                            row.append(item.assigned_to.get_full_name() if item.assigned_to else '')
                        else:
                            row.append(getattr(item, field, ''))
                else:
                    row = [
                        item.asset_tag or '',
                        item.name,
                        item.category.name if item.category else '',
                        item.status.name if item.status else '',
                        item.current_location.get_full_path() if item.current_location else '',
                        item.assigned_to.get_full_name() if item.assigned_to else '',
                        item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '',
                        item.purchase_cost or '',
                    ]
            writer.writerow(row)
        
        return response

    def export_excel(self, queryset, selected_fields=None):
        """Export to Excel format with field selection support"""
        if not EXCEL_AVAILABLE:
            messages.error(self.request, _("Excel ექსპორტი არ არის ხელმისაწვდომი. გთხოვთ დააინსტალიროთ openpyxl."))
            return redirect('inventory:equipment_list')
        
        wb = Workbook()
        ws = wb.active
        ws.title = _("აღჭურვილობის სია")
        
        # Define field mappings for headers
        field_headers = {
            'asset_tag': _('ინვენტარის ნომერი'),
            'name': _('დასახელება'),
            'category': _('კატეგორია'),
            'status': _('სტატუსი'),
            'location': _('მდებარეობა'),
            'purchase_date': _('შეძენის თარიღი'),
            'purchase_cost': _('შეძენის ფასი'),
            'assigned_to': _('მინიჭებული'),
        }
        
        # Write headers
        if selected_fields:
            headers = [field_headers[f] for f in selected_fields]
        else:
            headers = list(field_headers.values())
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            # Add header styling
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="C9C9C9", end_color="C9C9C9", fill_type="solid")
        
        # Write data rows
        for row_idx, item in enumerate(queryset, 2):
            if isinstance(item, dict):  # For custom field selection
                for col_idx, field in enumerate(selected_fields or field_headers.keys(), 1):
                    ws.cell(row=row_idx, column=col_idx, value=item.get(field, ''))
            else:  # For full equipment objects
                col_idx = 1
                if selected_fields:
                    for field in selected_fields:
                        value = ''
                        if field == 'category':
                            value = item.category.name if item.category else ''
                        elif field == 'status':
                            value = item.status.name if item.status else ''
                        elif field == 'location':
                            value = item.current_location.get_full_path() if item.current_location else ''
                        elif field == 'assigned_to':
                            value = item.assigned_to.get_full_name() if item.assigned_to else ''
                        elif field == 'purchase_date':
                            value = item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else ''
                        else:
                            value = getattr(item, field, '')
                        ws.cell(row=row_idx, column=col_idx, value=value)
                        col_idx += 1
                else:
                    ws.cell(row=row_idx, column=1, value=item.asset_tag or '')
                    ws.cell(row=row_idx, column=2, value=item.name)
                    ws.cell(row=row_idx, column=3, value=item.category.name if item.category else '')
                    ws.cell(row=row_idx, column=4, value=item.status.name if item.status else '')
                    ws.cell(row=row_idx, column=5, value=item.current_location.get_full_path() if item.current_location else '')
                    ws.cell(row=row_idx, column=6, value=item.assigned_to.get_full_name() if item.assigned_to else '')
                    ws.cell(row=row_idx, column=7, value=item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '')
                    ws.cell(row=row_idx, column=8, value=item.purchase_cost or '')
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="equipment_list.xlsx"'
        wb.save(response)
        return response

    def export_pdf(self, queryset, selected_fields=None):
        """Export to PDF format with field selection support"""
        if not PDF_AVAILABLE:
            messages.error(self.request, _("PDF ექსპორტი არ არის ხელმისაწვდომი. გთხოვთ დააინსტალიროთ reportlab."))
            return redirect('inventory:equipment_list')
        
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="equipment_list.pdf"'
        
        # Create PDF document
        doc = SimpleDocTemplate(response, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title = Paragraph(_("აღჭურვილობის ანგარიში"), styles['Title'])
        story.append(title)
        story.append(Spacer(1, 12))
        
        # Define field mappings for headers
        field_headers = {
            'asset_tag': _('ინვენტარის ნომერი'),
            'name': _('დასახელება'),
            'category': _('კატეგორია'),
            'status': _('სტატუსი'),
            'location': _('მდებარეობა'),
            'purchase_date': _('შეძენის თარიღი'),
            'purchase_cost': _('შეძენის ფასი'),
            'assigned_to': _('მინიჭებული'),
        }
        
        # Prepare table data
        if selected_fields:
            headers = [field_headers[f] for f in selected_fields]
        else:
            headers = list(field_headers.values())
        
        data = [headers]  # Start with headers
        
        # Add data rows
        for item in queryset:
            if isinstance(item, dict):  # For custom field selection
                row = [str(item.get(f, '')) for f in (selected_fields or field_headers.keys())]
            else:  # For full equipment objects
                if selected_fields:
                    row = []
                    for field in selected_fields:
                        if field == 'category':
                            row.append(item.category.name if item.category else '')
                        elif field == 'status':
                            row.append(item.status.name if item.status else '')
                        elif field == 'location':
                            row.append(item.current_location.get_full_path() if item.current_location else '')
                        elif field == 'assigned_to':
                            row.append(item.assigned_to.get_full_name() if item.assigned_to else '')
                        elif field == 'purchase_date':
                            row.append(item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '')
                        else:
                            row.append(str(getattr(item, field, '')))
                else:
                    row = [
                        item.asset_tag or '',
                        item.name[:30] + '...' if len(item.name) > 30 else item.name,
                        item.category.name[:20] if item.category else '',
                        item.status.name[:15] if item.status else '',
                        item.current_location.name[:20] if item.current_location else '',
                        item.assigned_to.get_full_name()[:20] if item.assigned_to else '',
                        item.purchase_date.strftime('%Y-%m-%d') if item.purchase_date else '',
                        str(item.purchase_cost) if item.purchase_cost else '',
                    ]
            data.append(row)
        
        # Create and style the table
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        # Add table to story
        story.append(table)
        
        # Add export information
        story.append(Spacer(1, 20))
        export_info = [
            Paragraph(f"{_('ექსპორტის თარიღი')}: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles['Normal']),
            Paragraph(f"{_('მთლიანი რაოდენობა')}: {len(data) - 1}", styles['Normal']),
        ]
        for info in export_info:
            story.append(info)
            story.append(Spacer(1, 6))
        
        # Build PDF
        doc.build(story)
        return response

    def get_queryset(self):
        queryset = super().get_queryset().select_related(
            'category', 'status', 'current_location', 'assigned_to', 'supplier'
        )
        
        status_filter_value = self.request.GET.get('status_filter')

        if status_filter_value == 'marked':
            marked_status = get_status_instance('Marked for Write-Off', is_marked_flag=True, is_decommissioned_flag=False)
            if marked_status:
                queryset = queryset.filter(status__is_marked_for_write_off=True, status__is_decommissioned=False)
            else: 
                queryset = queryset.none()
                messages.warning(self.request, _("The 'Marked for Write-Off' status is not configured correctly or no items match."))
        elif status_filter_value == 'decommissioned':
            decommissioned_status = get_status_instance('Decommissioned', is_decommissioned_flag=True)
            if decommissioned_status:
                 queryset = queryset.filter(status__is_decommissioned=True)
            else:
                queryset = queryset.none()
                messages.warning(self.request, _("The 'Decommissioned' status is not configured correctly or no items match."))
        elif status_filter_value and status_filter_value.isdigit():
            queryset = queryset.filter(status_id=int(status_filter_value))
        elif not status_filter_value or status_filter_value == 'active': 
            queryset = queryset.filter(status__is_decommissioned=False, status__is_marked_for_write_off=False)
        
        category_id_str = self.request.GET.get('category')
        if category_id_str and category_id_str.isdigit():
            queryset = queryset.filter(category_id=int(category_id_str))
        
        search_query = self.request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(asset_tag__icontains=search_query) |
                Q(serial_number__icontains=search_query) |
                Q(notes__icontains=search_query)
            )
            
        return queryset.order_by('name')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Equipment Inventory")
        context['categories_for_filter'] = Category.objects.all().order_by('name')
        context['statuses_for_filter'] = Status.objects.all().order_by('name')
        context['current_status_filter'] = self.request.GET.get('status_filter', 'active')
        
        context['status_active_filter_value'] = 'active'
        context['status_marked_filter_value'] = 'marked'
        context['status_decommissioned_filter_value'] = 'decommissioned'
        
        # For bulk operations - status and location lists
        context['status_list'] = Status.objects.filter(
            is_decommissioned=False, 
            is_marked_for_write_off=False
        ).order_by('name')
        context['location_list'] = Location.objects.all().order_by('name')
        
        # Equipment statistics for the enhanced dashboard
        total_equipment = Equipment.objects.all()
        context['active_count'] = total_equipment.filter(
            status__is_decommissioned=False, 
            status__is_marked_for_write_off=False
        ).count()
        context['maintenance_count'] = total_equipment.filter(
            status__is_marked_for_write_off=True,
            status__is_decommissioned=False
        ).count()
        context['decommissioned_count'] = total_equipment.filter(
            status__is_decommissioned=True
        ).count()
        
        # Count active filters for badge display
        active_filters = 0
        if self.request.GET.get('q'):
            active_filters += 1
        if self.request.GET.get('category'):
            active_filters += 1
        if self.request.GET.get('status_filter'):
            active_filters += 1
        context['active_filters_count'] = active_filters
        
        # Permissions used to show/hide links in template
        # These should align with permissions required by the target views
        context['can_add_equipment'] = self.request.user.has_perm('inventory.add_equipment')
        context['can_view_marked_list'] = self.request.user.has_perm('inventory.view_equipment') # Assuming MarkedForWriteOffListView requires this
        context['can_view_decommissioned_list'] = self.request.user.has_perm('inventory.view_equipment') # Assuming DecommissionedEquipmentListView requires this
        return context

class MarkedForWriteOffListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Equipment
    template_name = 'inventory/marked_for_write_off_list.html'
    context_object_name = 'equipment_list'
    permission_required = 'inventory.view_equipment' # Or a more specific perm like 'inventory.view_marked_equipment'
    paginate_by = 15

    def get_queryset(self):
        queryset = Equipment.objects.filter(
            status__is_marked_for_write_off=True, 
            status__is_decommissioned=False
        ).select_related(
            'category', 'status', 'current_location', 'assigned_to', 'supplier'
        ).order_by('-last_updated')
        
        search_query = self.request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(asset_tag__icontains=search_query) |
                Q(serial_number__icontains=search_query)
            )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Equipment Marked for Write-Off")
        context['list_type_description'] = _("This list shows equipment items that have been flagged for future decommissioning and are awaiting final action.")
        context['can_add_equipment'] = self.request.user.has_perm('inventory.add_equipment') # For consistency if you have a general "add" button
        return context

class DecommissionedEquipmentListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Equipment
    template_name = 'inventory/decommissioned_equipment_list.html'
    context_object_name = 'equipment_list'
    permission_required = 'inventory.view_equipment' # Or 'inventory.view_decommissioned_equipment'
    paginate_by = 15

    def get_queryset(self):
        queryset = Equipment.objects.filter(
            status__is_decommissioned=True
        ).select_related(
            'category', 'status', 'current_location', 'assigned_to', 'supplier', 'decommission_details'
        ).order_by('-decommission_details__decommission_date', '-last_updated')
        
        search_query = self.request.GET.get('q')
        if search_query:
            queryset = queryset.filter(
                Q(name__icontains=search_query) |
                Q(asset_tag__icontains=search_query) |
                Q(serial_number__icontains=search_query)
            )
        return queryset
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Decommissioned Equipment Archive")
        context['list_type_description'] = _("This list shows all equipment items that have been permanently decommissioned and removed from active service.")
        context['can_add_equipment'] = self.request.user.has_perm('inventory.add_equipment')
        return context


class EquipmentDetailView(LoginRequiredMixin, PermissionRequiredMixin, DetailView): # Added PermissionRequiredMixin
    model = Equipment
    template_name = 'inventory/equipment_detail.html'
    context_object_name = 'equipment'
    permission_required = 'inventory.view_equipment' # Added permission

    def get_queryset(self):
        return super().get_queryset().select_related(
            'category', 'status', 'current_location', 'assigned_to', 
            'supplier', 'added_by', 'updated_by'
        ).prefetch_related('decommission_details')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        equipment = self.object
        context['page_title'] = f"{equipment.name} ({_('Details')})"
        
        can_mark_for_write_off = False
        can_decommission = False
        can_restore = False

        # Check permissions for actions
        perm_change = self.request.user.has_perm('inventory.change_equipment')
        # Using 'delete_equipment' for decommission as per your current setup, 
        # but a specific 'decommission_equipment' perm would be better.
        perm_decommission = self.request.user.has_perm('inventory.delete_equipment') 
        # Using 'add_equipment' for restore as per your current setup,
        # but a specific 'restore_equipment' perm would be better.
        perm_restore = self.request.user.has_perm('inventory.add_equipment')

        if equipment.status:
            if not equipment.status.is_decommissioned and not equipment.status.is_marked_for_write_off:
                can_mark_for_write_off = perm_change
            if equipment.status.is_marked_for_write_off and not equipment.status.is_decommissioned:
                can_decommission = perm_decommission
            if equipment.status.is_decommissioned:
                can_restore = perm_restore
        
        context['can_edit_equipment'] = perm_change # For general edit button
        context['can_delete_equipment'] = self.request.user.has_perm('inventory.delete_equipment') # For general delete button (if applicable)

        context['can_mark_for_write_off'] = can_mark_for_write_off
        context['can_decommission'] = can_decommission
        context['can_restore'] = can_restore
        
        return context

class EquipmentCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = Equipment
    form_class = EquipmentForm
    template_name = 'inventory/equipment_form.html' 
    permission_required = 'inventory.add_equipment'
    success_url = reverse_lazy('inventory:equipment_list')
    success_message = _("Equipment '%(name)s' was created successfully.")

    def form_valid(self, form):
        form.instance.save(user=self.request.user) 
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Add New Equipment")
        context['form_title'] = _("Create Equipment Item")
        context['submit_button_text'] = _("Create Equipment")
        return context

class EquipmentUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Equipment
    form_class = EquipmentForm 
    template_name = 'inventory/equipment_form.html'
    permission_required = 'inventory.change_equipment'
    success_message = _("Equipment '%(name)s' was updated successfully.")

    def form_valid(self, form):
        new_status = form.cleaned_data.get('status')
        if new_status and (new_status.is_decommissioned or new_status.is_marked_for_write_off):
            if self.object.status != new_status:
                messages.error(self.request, _("Cannot set 'Decommissioned' or 'Marked for Write-Off' status directly. Please use the dedicated processes."))
                return self.form_invalid(form)
        
        form.instance.save(user=self.request.user)
        return super().form_valid(form)

    def get_success_url(self):
        return reverse_lazy('inventory:equipment_detail', kwargs={'pk': self.object.pk})

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Update {equipment_name}").format(equipment_name=self.object.name)
        context['form_title'] = _("Editing Equipment: {equipment_name}").format(equipment_name=self.object.name)
        context['submit_button_text'] = _("Save Changes")
        return context

class EquipmentDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Equipment
    template_name = 'inventory/equipment_confirm_delete.html'
    permission_required = 'inventory.delete_equipment'
    success_url = reverse_lazy('inventory:equipment_list')
    context_object_name = 'equipment'

    def dispatch(self, request, *args, **kwargs):
        equipment = self.get_object()
        if equipment.status and equipment.status.is_decommissioned:
            messages.error(request, _("Decommissioned equipment cannot be deleted. It serves as an archive record."))
            return redirect(equipment.get_absolute_url())
        if equipment.status and equipment.status.is_marked_for_write_off:
            messages.error(request, _("Equipment marked for write-off should be decommissioned, not deleted directly."))
            return redirect(equipment.get_absolute_url())
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        equipment_name = self.get_object().name
        messages.success(self.request, _("Equipment '{equipment_name}' has been successfully deleted.").format(equipment_name=equipment_name))
        return super().post(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Confirm Delete: {equipment_name}").format(equipment_name=self.object.name)
        return context

class EquipmentMarkForWriteOffView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    template_name = 'inventory/equipment_mark_for_write_off_form.html'
    form_class = EquipmentMarkForWriteOffForm
    permission_required = 'inventory.change_equipment' # Consider 'inventory.mark_equipment_for_write_off'

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        self.equipment_instance = get_object_or_404(Equipment, pk=self.kwargs['pk'])

    def dispatch(self, request, *args, **kwargs):
        if not self.equipment_instance.status or \
           self.equipment_instance.status.is_decommissioned or \
           self.equipment_instance.status.is_marked_for_write_off:
            messages.error(request, _("This equipment cannot be marked for write-off as it's already decommissioned, marked, or has no status."))
            return redirect(self.equipment_instance.get_absolute_url())
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['equipment'] = self.equipment_instance
        context['page_title'] = _("Mark for Write-Off: {equipment_name}").format(equipment_name=self.equipment_instance.name)
        context['form_title'] = _("Reason for Marking for Write-Off")
        context['submit_button_text'] = _("Mark for Write-Off")
        return context

    def form_valid(self, form):
        reason = form.cleaned_data['write_off_reason']
        marked_status = get_status_instance('Marked for Write-Off', is_marked_flag=True, is_decommissioned_flag=False)
        if not marked_status:
            messages.error(self.request, _("Critical Error: The 'Marked for Write-Off' status is not configured correctly in the system."))
            return redirect(self.equipment_instance.get_absolute_url())

        timestamp = timezone.now().strftime("%Y-%m-%d %H:%M")
        user_performing_action = self.request.user.username
        note_prefix = _("Marked for write-off on %(date)s by %(user)s.") % {'date': timestamp, 'user': user_performing_action}
        new_note_entry = f"{note_prefix}\n{_('Reason:')} {reason}"
        
        if self.equipment_instance.notes:
            self.equipment_instance.notes = f"{self.equipment_instance.notes}\n\n{new_note_entry}"
        else:
            self.equipment_instance.notes = new_note_entry
        
        self.equipment_instance.status = marked_status
        self.equipment_instance.save(user=self.request.user)

        EquipmentLog.objects.create(
             equipment=self.equipment_instance,
             user=self.request.user,
             change_type='marked_for_write_off',
             notes=_("Marked for write-off. Reason: {reason}").format(reason=reason)
        )
        messages.success(self.request, _("Equipment '{item_name}' has been successfully marked for write-off.").format(item_name=self.equipment_instance.name))
        return redirect(self.get_success_url())

    def get_success_url(self):
        return self.equipment_instance.get_absolute_url()


class EquipmentDecommissionView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    template_name = 'inventory/equipment_decommission_form.html'
    form_class = DecommissionConfirmForm
    permission_required = 'inventory.delete_equipment' # Consider 'inventory.decommission_equipment'

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        self.equipment_instance = get_object_or_404(Equipment, pk=self.kwargs['pk'])

    def dispatch(self, request, *args, **kwargs):
        if not (self.equipment_instance.status and self.equipment_instance.status.is_marked_for_write_off and not self.equipment_instance.status.is_decommissioned):
            messages.error(request, _("This equipment cannot be decommissioned. It must be 'Marked for Write-Off' first and not already decommissioned."))
            return redirect(self.equipment_instance.get_absolute_url())
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['equipment_instance'] = self.equipment_instance
        try:
            kwargs['instance'] = DecommissionLog.objects.get(equipment=self.equipment_instance)
        except DecommissionLog.DoesNotExist:
            pass
        return kwargs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['equipment'] = self.equipment_instance
        context['page_title'] = _("Decommission Equipment: {equipment_name}").format(equipment_name=self.equipment_instance.name)
        context['form_title'] = _("Confirm Decommissioning and Record Details")
        context['submit_button_text'] = _("Confirm Decommission")
        return context

    def form_valid(self, form):
        decommissioned_status = get_status_instance('Decommissioned', is_decommissioned_flag=True)
        if not decommissioned_status:
            messages.error(self.request, _("Critical Error: The 'Decommissioned' status is not configured correctly."))
            return redirect(self.equipment_instance.get_absolute_url())
        
        self.equipment_instance._decommission_form_data = form.cleaned_data
        self.equipment_instance.status = decommissioned_status
        self.equipment_instance.save(user=self.request.user)

        EquipmentLog.objects.create(
            equipment=self.equipment_instance,
            user=self.request.user,
            change_type='decommissioned',
            notes=_("Equipment decommissioned. Reason: {reason}").format(reason=form.cleaned_data.get('reason', _('N/A')))
        )
        messages.success(self.request, _("Equipment '{item_name}' has been successfully decommissioned.").format(item_name=self.equipment_instance.name))
        return redirect(self.get_success_url())

    def get_success_url(self):
        return self.equipment_instance.get_absolute_url()

class EquipmentRestoreView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    template_name = 'inventory/equipment_restore_form.html'
    form_class = EquipmentRestoreForm
    permission_required = 'inventory.add_equipment' # Consider 'inventory.restore_equipment'

    def setup(self, request, *args, **kwargs):
        super().setup(request, *args, **kwargs)
        self.equipment_instance = get_object_or_404(Equipment, pk=self.kwargs['pk'])

    def dispatch(self, request, *args, **kwargs):
        if not (self.equipment_instance.status and self.equipment_instance.status.is_decommissioned):
            messages.error(request, _("This equipment is not decommissioned and cannot be restored."))
            return redirect(self.equipment_instance.get_absolute_url())
        return super().dispatch(request, *args, **kwargs)
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['equipment'] = self.equipment_instance
        context['page_title'] = _("Restore Decommissioned Equipment: {equipment_name}").format(equipment_name=self.equipment_instance.name)
        context['form_title'] = _("Restore Equipment to Active Service")
        context['submit_button_text'] = _("Restore Equipment")
        return context

    def form_valid(self, form):
        restoration_reason = form.cleaned_data['restoration_reason']
        new_status = form.cleaned_data['new_status']

        timestamp = timezone.now().strftime("%Y-%m-%d %H:%M")
        user_performing_action = self.request.user.username
        note_prefix = _("Restored to service on %(date)s by %(user)s.") % {'date': timestamp, 'user': user_performing_action}
        new_note_entry = f"{note_prefix}\n{_('Restoration Reason:')} {restoration_reason}"
        
        if self.equipment_instance.notes:
            self.equipment_instance.notes = f"{self.equipment_instance.notes}\n\n{new_note_entry}"
        else:
            self.equipment_instance.notes = new_note_entry
        
        self.equipment_instance.status = new_status
        self.equipment_instance.save(user=self.request.user)

        EquipmentLog.objects.create(
            equipment=self.equipment_instance,
            user=self.request.user,
            change_type='restored',
            notes=_("Equipment restored. Reason: {reason}. New status: {status_name}").format(
                reason=restoration_reason, status_name=new_status.name
            )
        )
        messages.success(self.request, _("Equipment '{item_name}' has been successfully restored to status '{status_name}'.").format(
            item_name=self.equipment_instance.name, status_name=new_status.name
        ))
        return redirect(self.get_success_url())

    def get_success_url(self):
        return self.equipment_instance.get_absolute_url()

# --- CRUD Views for Category, Status, Location, Supplier ---

class CategoryListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Category
    template_name = 'inventory/category_list.html'
    context_object_name = 'categories'
    permission_required = 'inventory.view_category'
    paginate_by = 10
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Equipment Categories")
        context['can_add_category'] = self.request.user.has_perm('inventory.add_category')
        return context

class CategoryCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/generic_form.html'
    permission_required = 'inventory.add_category'
    success_url = reverse_lazy('inventory:category_list')
    success_message = _("Category '%(name)s' created successfully.")
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Create New Category")
        context['form_title'] = _("Add Category")
        context['submit_button_text'] = _("Create Category")
        return context

class CategoryUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Category
    form_class = CategoryForm
    template_name = 'inventory/generic_form.html'
    permission_required = 'inventory.change_category'
    success_url = reverse_lazy('inventory:category_list')
    success_message = _("Category '%(name)s' updated successfully.")
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Update Category: {object_name}").format(object_name=self.object.name)
        context['form_title'] = _("Edit Category: {object_name}").format(object_name=self.object.name)
        context['submit_button_text'] = _("Save Changes")
        return context

class CategoryDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Category
    template_name = 'inventory/generic_confirm_delete.html'
    permission_required = 'inventory.delete_category'
    success_url = reverse_lazy('inventory:category_list')
    context_object_name = 'object_to_delete' 
    
    def post(self, request, *args, **kwargs):
        object_name = self.get_object().name
        if self.get_object().equipment_items.exists() or self.get_object().stock_items.exists():
            messages.error(request, _("Category '{object_name}' cannot be deleted because it is currently in use by equipment or stock items.").format(object_name=object_name))
            return redirect('inventory:category_list')
        messages.success(self.request, _("Category '{object_name}' has been deleted.").format(object_name=object_name))
        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Confirm Delete: {object_name}").format(object_name=self.object.name)
        context['type_of_object'] = _("Category")
        if self.object.equipment_items.exists() or self.object.stock_items.exists():
            context['deletion_warning'] = _("This category is in use and cannot be deleted.")
        return context

class StatusListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Status
    template_name = 'inventory/status_list.html' 
    context_object_name = 'statuses'
    permission_required = 'inventory.view_status'
    paginate_by = 10
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Equipment Statuses")
        context['can_add_status'] = self.request.user.has_perm('inventory.add_status')
        return context

class StatusCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = Status
    form_class = StatusForm
    template_name = 'inventory/generic_form.html'
    permission_required = 'inventory.add_status'
    success_url = reverse_lazy('inventory:status_list')
    success_message = _("Status '%(name)s' created successfully.")
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Create New Status")
        context['form_title'] = _("Add Status")
        context['submit_button_text'] = _("Create Status")
        return context

class StatusUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Status
    form_class = StatusForm
    template_name = 'inventory/generic_form.html'
    permission_required = 'inventory.change_status'
    success_url = reverse_lazy('inventory:status_list')
    success_message = _("Status '%(name)s' updated successfully.")
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Update Status: {object_name}").format(object_name=self.object.name)
        context['form_title'] = _("Edit Status: {object_name}").format(object_name=self.object.name)
        context['submit_button_text'] = _("Save Changes")
        return context

class StatusDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Status
    template_name = 'inventory/generic_confirm_delete.html'
    permission_required = 'inventory.delete_status'
    success_url = reverse_lazy('inventory:status_list')
    context_object_name = 'object_to_delete'

    def post(self, request, *args, **kwargs):
        object_name = self.get_object().name
        if self.get_object().equipment_items_with_status.exists(): # Assuming related_name is 'equipment_items_with_status'
            messages.error(request, _("Status '{object_name}' cannot be deleted because it is currently in use by equipment items.").format(object_name=object_name))
            return redirect('inventory:status_list')
        messages.success(self.request, _("Status '{object_name}' has been deleted.").format(object_name=object_name))
        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Confirm Delete: {object_name}").format(object_name=self.object.name)
        context['type_of_object'] = _("Status")
        if self.get_object().equipment_items_with_status.exists():
            context['deletion_warning'] = _("This status is in use and cannot be deleted.")
        return context

class LocationListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Location
    template_name = 'inventory/location_list.html' 
    context_object_name = 'locations'
    permission_required = 'inventory.view_location'
    paginate_by = 10
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Locations")
        context['can_add_location'] = self.request.user.has_perm('inventory.add_location')
        return context

class LocationCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = Location
    form_class = LocationForm
    template_name = 'inventory/generic_form.html'
    permission_required = 'inventory.add_location'
    success_url = reverse_lazy('inventory:location_list')
    success_message = _("Location '%(name)s' created successfully.")
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Create New Location")
        context['form_title'] = _("Add Location")
        context['submit_button_text'] = _("Create Location")
        return context

class LocationUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Location
    form_class = LocationForm
    template_name = 'inventory/generic_form.html'
    permission_required = 'inventory.change_location'
    success_url = reverse_lazy('inventory:location_list')
    success_message = _("Location '%(name)s' updated successfully.")
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Update Location: {object_name}").format(object_name=self.object.name)
        context['form_title'] = _("Edit Location: {object_name}").format(object_name=self.object.name)
        context['submit_button_text'] = _("Save Changes")
        return context

class LocationDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Location
    template_name = 'inventory/generic_confirm_delete.html'
    permission_required = 'inventory.delete_location'
    success_url = reverse_lazy('inventory:location_list')
    context_object_name = 'object_to_delete'

    def post(self, request, *args, **kwargs):
        object_name = self.get_object().name
        if self.get_object().equipment_at_location.exists() or \
           self.get_object().stock_items_at_location.exists() or \
           self.get_object().child_locations.exists():
            messages.error(request, _("Location '{object_name}' cannot be deleted because it is in use or has child locations.").format(object_name=object_name))
            return redirect('inventory:location_list')
        messages.success(self.request, _("Location '{object_name}' has been deleted.").format(object_name=object_name))
        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Confirm Delete: {object_name}").format(object_name=self.object.name)
        context['type_of_object'] = _("Location")
        if self.get_object().equipment_at_location.exists() or \
           self.get_object().stock_items_at_location.exists() or \
           self.get_object().child_locations.exists():
            context['deletion_warning'] = _("This location is in use or has child locations and cannot be deleted.")
        return context

class SupplierListView(LoginRequiredMixin, PermissionRequiredMixin, ListView):
    model = Supplier
    template_name = 'inventory/supplier_list.html'
    context_object_name = 'suppliers'
    permission_required = 'inventory.view_supplier'
    paginate_by = 10
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Suppliers")
        context['can_add_supplier'] = self.request.user.has_perm('inventory.add_supplier')
        return context

class SupplierCreateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, CreateView):
    model = Supplier
    form_class = SupplierForm
    template_name = 'inventory/generic_form.html'
    permission_required = 'inventory.add_supplier'
    success_url = reverse_lazy('inventory:supplier_list')
    success_message = _("Supplier '%(name)s' created successfully.")
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Create New Supplier")
        context['form_title'] = _("Add Supplier")
        context['submit_button_text'] = _("Create Supplier")
        return context

class SupplierUpdateView(LoginRequiredMixin, PermissionRequiredMixin, SuccessMessageMixin, UpdateView):
    model = Supplier
    form_class = SupplierForm
    template_name = 'inventory/generic_form.html'
    permission_required = 'inventory.change_supplier'
    success_url = reverse_lazy('inventory:supplier_list')
    success_message = _("Supplier '%(name)s' updated successfully.")
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Update Supplier: {object_name}").format(object_name=self.object.name)
        context['form_title'] = _("Edit Supplier: {object_name}").format(object_name=self.object.name)
        context['submit_button_text'] = _("Save Changes")
        return context

class SupplierDeleteView(LoginRequiredMixin, PermissionRequiredMixin, DeleteView):
    model = Supplier
    template_name = 'inventory/generic_confirm_delete.html'
    permission_required = 'inventory.delete_supplier'
    success_url = reverse_lazy('inventory:supplier_list')
    context_object_name = 'object_to_delete'
    
    def post(self, request, *args, **kwargs): 
        object_name = self.get_object().name
        if self.get_object().supplied_equipment.exists() or self.get_object().supplied_stock_items.exists():
            messages.error(request, _("Supplier '{object_name}' cannot be deleted because it is linked to equipment or stock items.").format(object_name=object_name))
            return redirect('inventory:supplier_list')
        messages.success(self.request, _("Supplier '{object_name}' has been deleted.").format(object_name=object_name))
        return super().post(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = _("Confirm Delete: {object_name}").format(object_name=self.object.name)
        context['type_of_object'] = _("Supplier")
        if self.get_object().supplied_equipment.exists() or self.get_object().supplied_stock_items.exists():
            context['deletion_warning'] = _("This supplier is linked to equipment or stock items and cannot be deleted.")
        return context


# =======================
# BULK OPERATIONS & EXPORT VIEWS
# =======================

class EquipmentPDFExportView(LoginRequiredMixin, PermissionRequiredMixin, DetailView):
    """Individual Equipment PDF Export View"""
    model = Equipment
    permission_required = 'inventory.view_equipment'
    
    def get(self, request, *args, **kwargs):
        equipment = self.get_object()
        
        if not PDF_AVAILABLE:
            messages.error(request, _("PDF export is not available. Please install reportlab."))
            return redirect(equipment.get_absolute_url())
        
        try:
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="equipment_{equipment.asset_tag or equipment.pk}_details.pdf"'
            
            # Create PDF document
            doc = SimpleDocTemplate(response, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title = Paragraph(f"<b>Equipment Details: {equipment.name}</b>", styles['Title'])
            story.append(title)
            story.append(Spacer(1, 12))
            
            # Equipment information table
            data = [
                ['Field', 'Value'],
                ['Asset Tag', equipment.asset_tag or 'N/A'],
                ['Name', equipment.name],
                ['Category', equipment.category.name if equipment.category else 'N/A'],
                ['Status', equipment.status.name if equipment.status else 'N/A'],
                ['Location', equipment.current_location.get_full_path() if equipment.current_location else 'N/A'],
                ['Assigned To', equipment.assigned_to.get_full_name() if equipment.assigned_to else 'N/A'],
                ['Serial Number', equipment.serial_number or 'N/A'],
                ['Supplier', equipment.supplier.name if equipment.supplier else 'N/A'],
                ['Purchase Date', equipment.purchase_date.strftime('%Y-%m-%d') if equipment.purchase_date else 'N/A'],
                ['Purchase Cost', f"${equipment.purchase_cost}" if equipment.purchase_cost else 'N/A'],
                ['Warranty Expiry', equipment.warranty_expiry_date.strftime('%Y-%m-%d') if equipment.warranty_expiry_date else 'N/A'],
                ['Date Added', equipment.date_added.strftime('%Y-%m-%d %H:%M') if equipment.date_added else 'N/A'],
                ['Last Updated', equipment.last_updated.strftime('%Y-%m-%d %H:%M') if equipment.last_updated else 'N/A'],
            ]
            
            if equipment.notes:
                data.append(['Notes', equipment.notes[:100] + '...' if len(equipment.notes) > 100 else equipment.notes])
            
            table = Table(data, colWidths=[2*inch, 4*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            
            story.append(table)
            doc.build(story)
            
            return response
            
        except Exception as e:
            messages.error(request, _("Error generating PDF: {error}").format(error=str(e)))
            return redirect(equipment.get_absolute_url())


class BulkStatusChangeView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """Bulk status change for multiple equipment items"""
    permission_required = 'inventory.change_equipment'
    
    def post(self, request, *args, **kwargs):
        selected_ids = request.POST.getlist('selected_ids')
        new_status_id = request.POST.get('new_status')
        
        if not selected_ids:
            return JsonResponse({'success': False, 'error': _('No equipment items selected.')})
        
        if not new_status_id:
            return JsonResponse({'success': False, 'error': _('No status selected.')})
        
        try:
            new_status = Status.objects.get(pk=new_status_id)
            
            # Prevent setting decommissioned or marked status through bulk operations
            if new_status.is_decommissioned or new_status.is_marked_for_write_off:
                return JsonResponse({
                    'success': False, 
                    'error': _('Cannot set decommissioned or marked for write-off status through bulk operations.')
                })
            
            updated_count = 0
            errors = []
            
            for equipment_id in selected_ids:
                try:
                    equipment = Equipment.objects.get(pk=equipment_id)
                    
                    # Skip if already decommissioned
                    if equipment.status and equipment.status.is_decommissioned:
                        errors.append(f"Equipment {equipment.name} is decommissioned and cannot be updated.")
                        continue
                    
                    old_status = equipment.status
                    equipment.status = new_status
                    equipment.save(user=request.user)
                    
                    # Log the change
                    EquipmentLog.objects.create(
                        equipment=equipment,
                        user=request.user,
                        change_type='status_changed',
                        field_changed='status',
                        old_value=old_status.name if old_status else 'N/A',
                        new_value=new_status.name,
                        notes=_("Status changed via bulk operation.")
                    )
                    
                    updated_count += 1
                    
                except Equipment.DoesNotExist:
                    errors.append(f"Equipment with ID {equipment_id} not found.")
                except Exception as e:
                    errors.append(f"Error updating equipment ID {equipment_id}: {str(e)}")
            
            response_data = {
                'success': True,
                'updated_count': updated_count,
                'message': _("Successfully updated {count} equipment items.").format(count=updated_count)
            }
            
            if errors:
                response_data['warnings'] = errors
            
            return JsonResponse(response_data)
            
        except Status.DoesNotExist:
            return JsonResponse({'success': False, 'error': _('Selected status not found.')})
        except Exception as e:
            return JsonResponse({'success': False, 'error': _('An error occurred: {error}').format(error=str(e))})


class BulkLocationChangeView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """Bulk location change for multiple equipment items"""
    permission_required = 'inventory.change_equipment'
    
    def post(self, request, *args, **kwargs):
        selected_ids = request.POST.getlist('selected_ids')
        new_location_id = request.POST.get('new_location')
        
        if not selected_ids:
            return JsonResponse({'success': False, 'error': _('No equipment items selected.')})
        
        try:
            new_location = None
            if new_location_id:
                new_location = Location.objects.get(pk=new_location_id)
            
            updated_count = 0
            errors = []
            
            for equipment_id in selected_ids:
                try:
                    equipment = Equipment.objects.get(pk=equipment_id)
                    
                    # Skip if decommissioned
                    if equipment.status and equipment.status.is_decommissioned:
                        errors.append(f"Equipment {equipment.name} is decommissioned and cannot be moved.")
                        continue
                    
                    old_location = equipment.current_location
                    equipment.current_location = new_location
                    equipment.save(user=request.user)
                    
                    # Log the change
                    EquipmentLog.objects.create(
                        equipment=equipment,
                        user=request.user,
                        change_type='location_changed',
                        field_changed='current_location',
                        old_value=old_location.get_full_path() if old_location else 'N/A',
                        new_value=new_location.get_full_path() if new_location else 'N/A',
                        notes=_("Location changed via bulk operation.")
                    )
                    
                    updated_count += 1
                    
                except Equipment.DoesNotExist:
                    errors.append(f"Equipment with ID {equipment_id} not found.")
                except Exception as e:
                    errors.append(f"Error updating equipment ID {equipment_id}: {str(e)}")
            
            response_data = {
                'success': True,
                'updated_count': updated_count,
                'message': _("Successfully updated {count} equipment items.").format(count=updated_count)
            }
            
            if errors:
                response_data['warnings'] = errors
            
            return JsonResponse(response_data)
            
        except Location.DoesNotExist:
            return JsonResponse({'success': False, 'error': _('Selected location not found.')})
        except Exception as e:
            return JsonResponse({'success': False, 'error': _('An error occurred: {error}').format(error=str(e))})


class BulkDeleteView(LoginRequiredMixin, PermissionRequiredMixin, FormView):
    """Bulk delete for multiple equipment items"""
    permission_required = 'inventory.delete_equipment'
    
    def post(self, request, *args, **kwargs):
        selected_ids = request.POST.getlist('selected_ids')
        
        if not selected_ids:
            return JsonResponse({'success': False, 'error': _('No equipment items selected.')})
        
        try:
            deleted_count = 0
            errors = []
            
            for equipment_id in selected_ids:
                try:
                    equipment = Equipment.objects.get(pk=equipment_id)
                    
                    # Prevent deletion of decommissioned or marked equipment
                    if equipment.status and (equipment.status.is_decommissioned or equipment.status.is_marked_for_write_off):
                        errors.append(f"Equipment {equipment.name} cannot be deleted (decommissioned/marked status).")
                        continue
                    
                    equipment_name = equipment.name
                    equipment.delete()
                    deleted_count += 1
                    
                except Equipment.DoesNotExist:
                    errors.append(f"Equipment with ID {equipment_id} not found.")
                except Exception as e:
                    errors.append(f"Error deleting equipment ID {equipment_id}: {str(e)}")
            
            response_data = {
                'success': True,
                'deleted_count': deleted_count,
                'message': _("Successfully deleted {count} equipment items.").format(count=deleted_count)
            }
            
            if errors:
                response_data['warnings'] = errors
            
            return JsonResponse(response_data)
            
        except Exception as e:
            return JsonResponse({'success': False, 'error': _('An error occurred: {error}').format(error=str(e))})
